import getpass
from pathlib import Path

import psycopg
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.worksheet.table import Table, TableStyleInfo


# =========================================================
# PROJECT SETTINGS
# =========================================================

PROJECT_ROOT = Path(__file__).resolve().parent.parent
REPORTS_DIR = PROJECT_ROOT / "reports"

REPORTS_DIR.mkdir(parents=True, exist_ok=True)

OUTPUT_FILE = (
    REPORTS_DIR
    / "customer_support_management_report.xlsx"
)


# =========================================================
# POSTGRESQL SETTINGS
# =========================================================

DB_NAME = "customer_support_analytics"
DB_USER = "postgres"
DB_HOST = "localhost"
DB_PORT = "5432"


# =========================================================
# REPORT STYLING
# =========================================================

TITLE_FILL = PatternFill(
    "solid",
    fgColor="1F4E78"
)

HEADER_FILL = PatternFill(
    "solid",
    fgColor="D9EAF7"
)

CRITICAL_FILL = PatternFill(
    "solid",
    fgColor="F4CCCC"
)

HIGH_FILL = PatternFill(
    "solid",
    fgColor="FCE5CD"
)

WARNING_FILL = PatternFill(
    "solid",
    fgColor="FFF2CC"
)

ON_TRACK_FILL = PatternFill(
    "solid",
    fgColor="D9EAD3"
)

GOOD_FILL = PatternFill(
    "solid",
    fgColor="D9EAD3"
)

POOR_FILL = PatternFill(
    "solid",
    fgColor="F4CCCC"
)

NEUTRAL_FILL = PatternFill(
    "solid",
    fgColor="FFF2CC"
)

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


# =========================================================
# SQL QUERIES
# =========================================================

KPI_QUERY = """
SELECT
    COUNT(*) AS total_tickets,

    COUNT(*) FILTER (
        WHERE status = 'Open'
    ) AS open_tickets,

    COUNT(*) FILTER (
        WHERE status = 'Closed'
    ) AS closed_tickets,

    COUNT(*) FILTER (
        WHERE status = 'Open'
        AND sla_breached = TRUE
    ) AS overdue_open_tickets,

    ROUND(
        AVG(resolution_hours)
        FILTER (
            WHERE status = 'Closed'
        ),
        2
    ) AS avg_resolution_hours,

    ROUND(
        100.0 *
        COUNT(*) FILTER (
            WHERE status = 'Closed'
            AND sla_breached = FALSE
        )
        /
        NULLIF(
            COUNT(*) FILTER (
                WHERE status = 'Closed'
            ),
            0
        ),
        2
    ) AS closed_sla_compliance_pct

FROM tickets;
"""


TEAM_QUERY = """
WITH ticket_metrics AS (

    SELECT
        a.team,

        COUNT(*) AS total_tickets,

        COUNT(*) FILTER (
            WHERE t.status = 'Open'
        ) AS open_tickets,

        COUNT(*) FILTER (
            WHERE t.status = 'Open'
            AND t.sla_breached = TRUE
        ) AS overdue_open_tickets,

        ROUND(
            100.0 *
            COUNT(*) FILTER (
                WHERE t.status = 'Closed'
                AND t.sla_breached = FALSE
            )
            /
            NULLIF(
                COUNT(*) FILTER (
                    WHERE t.status = 'Closed'
                ),
                0
            ),
            2
        ) AS sla_compliance_pct

    FROM tickets t

    JOIN agents a
        ON t.agent_id = a.agent_id

    GROUP BY a.team
),

qa_metrics AS (

    SELECT
        a.team,

        ROUND(
            AVG(q.qa_score),
            2
        ) AS avg_qa_score,

        ROUND(
            100.0 *
            COUNT(*) FILTER (
                WHERE q.passed_qa = TRUE
            )
            /
            NULLIF(
                COUNT(q.evaluation_id),
                0
            ),
            2
        ) AS qa_pass_rate_pct,

        COUNT(*) FILTER (
            WHERE q.critical_error = TRUE
        ) AS critical_errors

    FROM qa_evaluations q

    JOIN agents a
        ON q.agent_id = a.agent_id

    GROUP BY a.team
)

SELECT
    t.team,
    t.total_tickets,
    t.open_tickets,
    t.overdue_open_tickets,
    t.sla_compliance_pct,
    q.avg_qa_score,
    q.qa_pass_rate_pct,
    q.critical_errors

FROM ticket_metrics t

JOIN qa_metrics q
    ON t.team = q.team

ORDER BY
    t.overdue_open_tickets DESC;
"""


EXCEPTION_QUERY = """
SELECT
    t.ticket_id,
    a.agent_name,
    a.team,
    t.category,
    t.priority,
    t.opened_at,
    t.age_hours,
    t.sla_target_hours,

    GREATEST(
        t.age_hours - t.sla_target_hours,
        0
    ) AS hours_over_sla,

    CASE
        WHEN t.priority = 'Critical'
             AND t.sla_breached = TRUE
            THEN 'Critical'

        WHEN t.priority = 'High'
             AND t.sla_breached = TRUE
            THEN 'High'

        WHEN t.sla_breached = TRUE
             AND (
                 t.age_hours
                 - t.sla_target_hours
             ) >= 168
            THEN 'High'

        WHEN t.sla_breached = TRUE
            THEN 'Warning'

        ELSE 'On Track'
    END AS management_priority

FROM tickets t

JOIN agents a
    ON t.agent_id = a.agent_id

WHERE t.status = 'Open'

ORDER BY
    CASE
        WHEN t.priority = 'Critical'
             AND t.sla_breached = TRUE
            THEN 1

        WHEN t.priority = 'High'
             AND t.sla_breached = TRUE
            THEN 2

        WHEN t.sla_breached = TRUE
             AND (
                 t.age_hours
                 - t.sla_target_hours
             ) >= 168
            THEN 3

        WHEN t.sla_breached = TRUE
            THEN 4

        ELSE 5
    END,

    hours_over_sla DESC;
"""


# =========================================================
# HELPER FUNCTIONS
# =========================================================

def style_title(
    worksheet,
    title,
    end_column
):
    """
    Create a management-style title row.
    """

    worksheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=end_column
    )

    title_cell = worksheet.cell(
        row=1,
        column=1,
        value=title
    )

    title_cell.font = Font(
        bold=True,
        size=16,
        color="FFFFFF"
    )

    title_cell.fill = TITLE_FILL

    title_cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    worksheet.row_dimensions[1].height = 26


def style_header_row(
    worksheet,
    row_number,
    column_count
):
    """
    Apply consistent formatting to table headers.
    """

    for column_number in range(
        1,
        column_count + 1
    ):

        cell = worksheet.cell(
            row=row_number,
            column=column_number
        )

        cell.font = Font(
            bold=True
        )

        cell.fill = HEADER_FILL

        cell.border = THIN_BORDER

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )


def auto_size_columns(
    worksheet,
    minimum_width=10,
    maximum_width=30
):
    """
    Automatically size worksheet columns
    while safely ignoring merged cells.
    """

    from openpyxl.utils import get_column_letter

    for column_number, column_cells in enumerate(
        worksheet.iter_cols(),
        start=1
    ):

        max_length = 0

        for cell in column_cells:

            # Ignore cells created by merged ranges
            if isinstance(cell, MergedCell):
                continue

            if cell.value is not None:

                cell_length = len(
                    str(cell.value)
                )

                if cell_length > max_length:
                    max_length = cell_length

        column_letter = get_column_letter(
            column_number
        )

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            max(
                max_length + 2,
                minimum_width
            ),
            maximum_width
        )


def add_excel_table(
    worksheet,
    table_name,
    start_row,
    end_row,
    end_column
):
    """
    Turn a worksheet range into an Excel table.
    """

    from openpyxl.utils import (
        get_column_letter
    )

    end_letter = get_column_letter(
        end_column
    )

    table_reference = (
        f"A{start_row}:"
        f"{end_letter}{end_row}"
    )

    table = Table(
        displayName=table_name,
        ref=table_reference
    )

    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )

    table.tableStyleInfo = style

    worksheet.add_table(table)


# =========================================================
# GET POSTGRESQL PASSWORD
# =========================================================

password = getpass.getpass(
    "PostgreSQL password: "
)


# =========================================================
# QUERY POSTGRESQL
# =========================================================

with psycopg.connect(
    dbname=DB_NAME,
    user=DB_USER,
    password=password,
    host=DB_HOST,
    port=DB_PORT
) as connection:

    with connection.cursor() as cursor:

        print(
            "Connected to PostgreSQL."
        )

        # -------------------------
        # KPI data
        # -------------------------

        cursor.execute(
            KPI_QUERY
        )

        kpi_result = (
            cursor.fetchone()
        )

        # -------------------------
        # Team performance data
        # -------------------------

        cursor.execute(
            TEAM_QUERY
        )

        team_rows = (
            cursor.fetchall()
        )

        # -------------------------
        # Exception data
        # -------------------------

        cursor.execute(
            EXCEPTION_QUERY
        )

        exception_rows = (
            cursor.fetchall()
        )


# =========================================================
# CREATE WORKBOOK
# =========================================================

workbook = Workbook()


# =========================================================
# SHEET 1 — KPI SUMMARY
# =========================================================

ws_kpi = workbook.active

ws_kpi.title = "KPI Summary"


style_title(
    ws_kpi,
    "Customer Support Management KPIs",
    4
)


kpi_data = [
    (
        "Total Tickets",
        kpi_result[0]
    ),
    (
        "Open Tickets",
        kpi_result[1]
    ),
    (
        "Closed Tickets",
        kpi_result[2]
    ),
    (
        "Overdue Open Tickets",
        kpi_result[3]
    ),
    (
        "Average Resolution Hours",
        kpi_result[4]
    ),
    (
        "Closed SLA Compliance %",
        kpi_result[5]
    ),
]


for row_number, (
    label,
    value
) in enumerate(
    kpi_data,
    start=3
):

    label_cell = ws_kpi.cell(
        row=row_number,
        column=1,
        value=label
    )

    value_cell = ws_kpi.cell(
        row=row_number,
        column=2,
        value=value
    )

    label_cell.font = Font(
        bold=True
    )

    label_cell.fill = (
        HEADER_FILL
    )

    label_cell.border = (
        THIN_BORDER
    )

    value_cell.border = (
        THIN_BORDER
    )

    label_cell.alignment = Alignment(
        vertical="center"
    )

    value_cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )


# Number formats

ws_kpi["B7"].number_format = (
    "0.00"
)

ws_kpi["B8"].number_format = (
    '0.00"%"'
)


# Highlight key risk indicators

ws_kpi["B6"].fill = (
    WARNING_FILL
)

if float(kpi_result[5]) < 80:

    ws_kpi["B8"].fill = (
        WARNING_FILL
    )

else:

    ws_kpi["B8"].fill = (
        GOOD_FILL
    )


# KPI column sizes

ws_kpi.column_dimensions[
    "A"
].width = 30

ws_kpi.column_dimensions[
    "B"
].width = 18


# =========================================================
# SHEET 2 — TEAM PERFORMANCE
# =========================================================

ws_team = workbook.create_sheet(
    "Team Performance"
)


style_title(
    ws_team,
    "Team Performance Overview",
    8
)


team_headers = [
    "Team",
    "Total Tickets",
    "Open Tickets",
    "Overdue Open Tickets",
    "SLA Compliance %",
    "Average QA Score",
    "QA Pass Rate %",
    "Critical Errors"
]


for column_number, header in enumerate(
    team_headers,
    start=1
):

    ws_team.cell(
        row=3,
        column=column_number,
        value=header
    )


style_header_row(
    ws_team,
    3,
    len(team_headers)
)


for row_number, row_data in enumerate(
    team_rows,
    start=4
):

    for column_number, value in enumerate(
        row_data,
        start=1
    ):

        cell = ws_team.cell(
            row=row_number,
            column=column_number,
            value=value
        )

        cell.border = THIN_BORDER

        cell.alignment = Alignment(
            vertical="center"
        )


# Percentage and score formatting

for row_number in range(
    4,
    4 + len(team_rows)
):

    ws_team.cell(
        row=row_number,
        column=5
    ).number_format = (
        '0.00"%"'
    )

    ws_team.cell(
        row=row_number,
        column=6
    ).number_format = (
        "0.00"
    )

    ws_team.cell(
        row=row_number,
        column=7
    ).number_format = (
        '0.00"%"'
    )


# Conditional visual emphasis

for row_number in range(
    4,
    4 + len(team_rows)
):

    sla_cell = ws_team.cell(
        row=row_number,
        column=5
    )

    qa_cell = ws_team.cell(
        row=row_number,
        column=7
    )

    critical_cell = ws_team.cell(
        row=row_number,
        column=8
    )

    # SLA performance

    if sla_cell.value < 70:

        sla_cell.fill = (
            POOR_FILL
        )

    elif sla_cell.value < 80:

        sla_cell.fill = (
            NEUTRAL_FILL
        )

    else:

        sla_cell.fill = (
            GOOD_FILL
        )

    # QA pass rate

    if qa_cell.value < 75:

        qa_cell.fill = (
            POOR_FILL
        )

    elif qa_cell.value < 85:

        qa_cell.fill = (
            NEUTRAL_FILL
        )

    else:

        qa_cell.fill = (
            GOOD_FILL
        )

    # Critical errors

    if critical_cell.value > 0:

        critical_cell.fill = (
            CRITICAL_FILL
        )


# Freeze headers

ws_team.freeze_panes = "A4"


# Excel table

if team_rows:

    add_excel_table(
        ws_team,
        "TeamPerformanceTable",
        3,
        3 + len(team_rows),
        len(team_headers)
    )


auto_size_columns(
    ws_team
)


# =========================================================
# SHEET 3 — EXCEPTION REGISTER
# =========================================================

ws_exceptions = workbook.create_sheet(
    "Exception Register"
)


style_title(
    ws_exceptions,
    "Prioritized Open-Ticket Exception Register",
    10
)


exception_headers = [
    "Ticket ID",
    "Agent Name",
    "Team",
    "Category",
    "Priority",
    "Opened At",
    "Age Hours",
    "SLA Target Hours",
    "Hours Over SLA",
    "Management Priority"
]


for column_number, header in enumerate(
    exception_headers,
    start=1
):

    ws_exceptions.cell(
        row=3,
        column=column_number,
        value=header
    )


style_header_row(
    ws_exceptions,
    3,
    len(exception_headers)
)


for row_number, row_data in enumerate(
    exception_rows,
    start=4
):

    for column_number, value in enumerate(
        row_data,
        start=1
    ):

        cell = ws_exceptions.cell(
            row=row_number,
            column=column_number,
            value=value
        )

        cell.border = THIN_BORDER

        cell.alignment = Alignment(
            vertical="center"
        )


# Date formatting

for row_number in range(
    4,
    4 + len(exception_rows)
):

    ws_exceptions.cell(
        row=row_number,
        column=6
    ).number_format = (
        "yyyy-mm-dd hh:mm"
    )


# Highlight management priority

for row_number in range(
    4,
    4 + len(exception_rows)
):

    priority_cell = (
        ws_exceptions.cell(
            row=row_number,
            column=10
        )
    )

    priority_value = (
        priority_cell.value
    )

    if priority_value == "Critical":

        priority_cell.fill = (
            CRITICAL_FILL
        )

        priority_cell.font = Font(
            bold=True
        )

    elif priority_value == "High":

        priority_cell.fill = (
            HIGH_FILL
        )

        priority_cell.font = Font(
            bold=True
        )

    elif priority_value == "Warning":

        priority_cell.fill = (
            WARNING_FILL
        )

    elif priority_value == "On Track":

        priority_cell.fill = (
            ON_TRACK_FILL
        )


# Freeze header row

ws_exceptions.freeze_panes = "A4"


# Excel table

if exception_rows:

    add_excel_table(
        ws_exceptions,
        "ExceptionRegisterTable",
        3,
        3 + len(exception_rows),
        len(exception_headers)
    )


auto_size_columns(
    ws_exceptions
)


# =========================================================
# GENERAL WORKBOOK FORMATTING
# =========================================================

for worksheet in workbook.worksheets:

    worksheet.sheet_view.showGridLines = False

    worksheet.page_setup.orientation = (
        "landscape"
    )

    worksheet.page_setup.fitToWidth = 1

    worksheet.page_setup.fitToHeight = 0


# =========================================================
# SAVE REPORT
# =========================================================

try:

    workbook.save(
        OUTPUT_FILE
    )

except PermissionError:

    print()
    print(
        "Unable to save the Excel report."
    )

    print(
        "Please close the existing report "
        "in Excel and run the script again."
    )

    raise


print()
print(
    "Excel management report created successfully."
)

print(
    f"Saved to: {OUTPUT_FILE}"
)